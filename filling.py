import openpyxl

def create_labyrinth_excel(labyrinth_text, output_filename='labyrinth.xlsx'):
    # Convert the text input to a 2D array
    labyrinth_array = [list(map(int, line.split())) for line in labyrinth_text.strip().split('\n')]

    # Create Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row_index, row in enumerate(labyrinth_array):
        for col_index, cell_value in enumerate(row):
            cell = sheet.cell(row=row_index + 1, column=col_index + 1)
            if cell_value == 0:
                # Set black color for gates
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            elif cell_value == 1:
                # Set white color for walls
                cell.fill = openpyxl.styles.PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    workbook.save(output_filename)

# Example text input
labyrinth_text = """
0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0
1 1 1 1 1 1 1 1 1 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0
0 0 0 0 0 0 0 0 1 0 1 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 0
0 1 1 1 1 1 1 0 1 0 0 1 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0 1 0
0 1 0 0 0 0 0 0 1 1 0 0 0 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 0
0 1 0 1 1 1 0 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0 1 1 1 0 1 0
0 1 0 1 0 0 1 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0 1 0 1 0 1 0
0 1 0 1 0 1 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 0 0 0 1 0 1 0
0 1 0 1 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0 0 1 0
0 1 0 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 1 0
0 1 0 1 0 1 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0
0 1 0 1 0 1 0 1 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 0
0 1 0 1 0 1 0 1 1 0 1 1 1 1 1 1 1 1 0 1 1 1 1 1 1 1 1 0 1 0
0 0 0 1 0 1 0 1 1 0 1 1 0 0 0 0 0 1 0 1 0 0 0 0 0 0 1 0 1 0
0 1 0 1 0 1 0 1 1 0 1 1 0 0 0 0 0 1 0 1 0 1 1 1 1 0 1 0 1 0
0 1 0 1 0 1 0 1 1 0 1 1 0 0 0 0 0 1 0 1 0 1 0 0 0 0 1 0 1 0
0 1 0 1 0 1 0 1 1 0 1 1 0 0 0 0 0 1 0 1 0 1 0 1 1 1 1 0 1 0
0 1 1 1 0 1 1 1 1 0 1 0 0 0 0 0 0 1 1 1 0 1 0 1 0 1 1 0 1 0
0 0 0 0 0 0 0 0 0 0 1 0 1 1 1 1 1 0 0 0 0 1 0 1 0 1 1 0 1 0
0 0 1 1 1 1 1 1 1 1 1 0 1 1 0 1 1 0 1 0 1 1 0 1 0 1 1 0 1 0
0 1 0 0 0 0 0 0 0 0 0 0 1 0 0 0 0 1 1 0 1 1 0 1 0 1 1 0 1 0
0 1 1 1 1 1 1 1 1 1 1 1 1 0 1 1 0 0 0 0 1 1 0 0 0 1 1 0 1 0
0 1 0 1 1 1 1 1 1 1 0 1 1 1 1 1 0 1 1 1 1 1 0 1 1 1 1 0 1 0
0 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 1 0 1 1 0 0 0 0 1 0 1 0
0 1 1 1 1 1 1 1 1 0 1 1 1 1 1 1 1 1 1 0 1 1 0 1 1 0 1 0 1 0
0 1 0 0 0 0 0 0 0 0 1 0 0 0 0 0 0 0 0 0 0 0 0 1 1 0 1 0 1 0
0 1 0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0 1 0
0 1 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 1 0
0 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 1 0 1
0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0 0
"""

create_labyrinth_excel(labyrinth_text)
